import streamlit as st
import pandas as pd
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import numpy as np
import io
import plotly.graph_objects as go

# Title of the app
st.title("Stand Ops Report Generator")

# File uploader
uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

if uploaded_file is not None:
    # Read the CSV file
    df = pd.read_csv(uploaded_file)
    
    # Display the uploaded CSV
    st.write("Uploaded CSV file:")
    st.dataframe(df)

    # Process the DataFrame
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%y").dt.date
    df["Todays Date"] = dt.datetime.now().date()
    df["To Ship On"] = pd.to_datetime(df["To Ship On"], format="%m/%d/%y").dt.date
    df["Days Since Order Received"] = (df["Todays Date"] - df["Date"]).dt.days
    df["Lead Time"] = (df["Todays Date"] - df["To Ship On"]).dt.days
    df = df[['Date', 'Todays Date', 'To Ship On', 'Days Since Order Received',
             'Lead Time', 'Document Number', 'PO Number', 'Customer', 'Status', 'Open Amount']]
    df['sample_lead_times'] = np.random.randint(1, 25, size=len(df))

    # Adding a color column based on conditional formatting rules
    df['color'] = np.where(df['sample_lead_times'] >= df['Lead Time'], 'blue', 'red')

    # Visualize the DataFrame with conditional colors using Plotly
    fig = go.Figure(data=[go.Table(
        header=dict(values=list(df.columns[:-1]), # Exclude 'color' from header
                    fill_color='paleturquoise',
                    align='left'),
        cells=dict(values=[df[col] for col in df.columns[:-1]], # Exclude 'color' from cells
                   fill_color=[[color for color in df['color']]],
                   align='left'))
    ])

    st.plotly_chart(fig, use_container_width=True)

    # Write the DataFrame to an Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        writer.save()

    # Load the Excel file using openpyxl
    output.seek(0)
    workbook = load_workbook(output)
    sheet = workbook.active

    # Get the column letters for 'Lead Time' and 'sample_lead_times'
    lead_times_col = df.columns.get_loc('Lead Time') + 1  # +1 because openpyxl is 1-based index
    sample_lead_times_col = df.columns.get_loc('sample_lead_times') + 1

    # Define the conditional formatting rules
    blue_fill = PatternFill(start_color='8BD3E6', end_color='8BD3E6', fill_type='solid')
    red_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')

    # Apply conditional formatting to each row
    for row in range(2, sheet.max_row + 1):  # Start from 2 to skip the header row
        lead_times_cell = sheet.cell(row=row, column=lead_times_col).coordinate
        sample_lead_times_cell = sheet.cell(row=row, column=sample_lead_times_col).coordinate
        sheet.conditional_formatting.add(f"A{row}:Z{row}",
            FormulaRule(
                formula=[f"${sample_lead_times_cell} >= ${lead_times_cell}"],
                fill=blue_fill
            )
        )
        sheet.conditional_formatting.add(f"A{row}:Z{row}",
            FormulaRule(
                formula=[f"${sample_lead_times_cell} < ${lead_times_cell}"],
                fill=red_fill
            )
        )

    # Save the workbook to a BytesIO object
    processed_file = io.BytesIO()
    workbook.save(processed_file)
    processed_file.seek(0)

    # Generate filename with the current date
    current_date = dt.datetime.now().strftime("%Y-%m-%d")
    filename = f'Taylor_Report_{current_date}.xlsx'

    # Provide a link to download the processed Excel file
    st.download_button(
        label="Download processed Excel file",
        data=processed_file,
        file_name=filename,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
