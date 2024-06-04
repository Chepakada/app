import pandas as pd
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import numpy as np

df = pd.read_csv("C:/Users/19518/Downloads/taylor_data.csv")

df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%y").dt.date
df["Todays Date"] = dt.datetime.now().date()
df["To Ship On"] = pd.to_datetime(df["To Ship On"], format="%m/%d/%y").dt.date

df["Days Since Order Received"] = (df["Todays Date"] - df["Date"]).dt.days
df["Lead Time"] = (df["Todays Date"] - df["To Ship On"]).dt.days

df = df[['Date', 'Todays Date', 'To Ship On', 'Days Since Order Received',
         'Lead Time', 'Document Number', 'PO Number', 'Customer', 'Status', 'Open Amount']]

df['sample_lead_times'] = np.random.randint(1, 25, size=len(df))
# Step 2: Write the DataFrame to an Excel file
excel_file_path = 'SampleFile.xlsx'
df.to_excel(excel_file_path, index=False)
# Step 3: Load the Excel file using openpyxl
workbook = load_workbook(excel_file_path)
sheet = workbook.active
# Get the column letters for 'Lead Times' and 'sample_lead_times'
lead_times_col = df.columns.get_loc('Lead Time') + 1  # +1 because openpyxl is 1-based index
sample_lead_times_col = df.columns.get_loc('sample_lead_times') + 1
# Define the conditional formatting rules
blue_fill = PatternFill(start_color='8BD3E6', end_color='8BD3E6', fill_type='solid')
red_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
# Apply conditional formatting to each row
for row in range(2, sheet.max_row + 1):  # Start from 2 to skip the header row
    lead_times_cell = sheet.cell(row=row, column=lead_times_col).coordinate
    sample_lead_times_cell = sheet.cell(row=row, column=sample_lead_times_col).coordinate
    sheet.conditional_formatting.add(f"A{row}:Z{row}",
        FormulaRule(
            formula=[f"${sample_lead_times_cell} >= ${lead_times_cell}"],
            fill=blue_fill
        )
    )
    sheet.conditional_formatting.add(f"A{row}:Z{row}",
        FormulaRule(
            formula=[f"${sample_lead_times_cell} < ${lead_times_cell}"],
            fill=red_fill
        )
    )
# Save the workbook
workbook.save(excel_file_path)
